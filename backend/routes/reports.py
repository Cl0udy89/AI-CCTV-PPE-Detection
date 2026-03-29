"""Report generation: PDF (reportlab) and Excel (openpyxl)."""
import io, os, sqlite3
from datetime import datetime
from fastapi import APIRouter, Query
from fastapi.responses import StreamingResponse
from core.auth import require_role

router = APIRouter(prefix="/reports", tags=["reports"])

INC_DB = os.path.join(os.path.dirname(__file__), '..', 'data', 'incidents.db')

# ── Translations ──────────────────────────────────────────────────────────────

_T = {
    'pl': {
        'title_default':   'Raport incydentów PPE',
        'generated':       'Wygenerowano',
        'period':          'Okres',
        'all':             'wszystkie',
        'summary_total':   'Łącznie',
        'summary_new':     'Nowe',
        'summary_closed':  'Zamknięte',
        'summary_review':  'W trakcie',
        'no_incidents':    'Brak incydentów w wybranym okresie.',
        'col_id':          '#',
        'col_date':        'Data',
        'col_violation':   'Typ naruszenia',
        'col_zone':        'Strefa',
        'col_status':      'Status',
        'col_track':       'Track ID',
        'col_notes':       'Notatki',
        'sheet_incidents': 'Incydenty PPE',
        'sheet_summary':   'Podsumowanie',
        'status_new':      'nowy',
        'status_review':   'w trakcie',
        'status_closed':   'zamknięty',
        'violations': {
            'NO-Hardhat':     'Brak kasku',
            'NO-Safety Vest': 'Brak kamizelki',
            'NO-Mask':        'Brak maski',
        },
    },
    'en': {
        'title_default':   'PPE Incidents Report',
        'generated':       'Generated',
        'period':          'Period',
        'all':             'all',
        'summary_total':   'Total',
        'summary_new':     'New',
        'summary_closed':  'Closed',
        'summary_review':  'Reviewing',
        'no_incidents':    'No incidents in the selected period.',
        'col_id':          '#',
        'col_date':        'Date',
        'col_violation':   'Violation type',
        'col_zone':        'Zone',
        'col_status':      'Status',
        'col_track':       'Track ID',
        'col_notes':       'Notes',
        'sheet_incidents': 'PPE Incidents',
        'sheet_summary':   'Summary',
        'status_new':      'new',
        'status_review':   'reviewing',
        'status_closed':   'closed',
        'violations': {
            'NO-Hardhat':     'No hardhat',
            'NO-Safety Vest': 'No safety vest',
            'NO-Mask':        'No mask',
        },
    },
}


def _tr(lang: str) -> dict:
    return _T.get(lang, _T['pl'])


# ── Unicode font registration ─────────────────────────────────────────────────

_FONT_REGULAR_PATHS = [
    '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    '/usr/share/fonts/dejavu/DejaVuSans.ttf',
    '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
    '/usr/share/fonts/liberation/LiberationSans-Regular.ttf',
    'C:/Windows/Fonts/arial.ttf',
    'C:/Windows/Fonts/calibri.ttf',
    '/Library/Fonts/Arial.ttf',
    '/System/Library/Fonts/Supplemental/Arial.ttf',
]

_FONT_BOLD_PATHS = [
    '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
    '/usr/share/fonts/dejavu/DejaVuSans-Bold.ttf',
    '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf',
    '/usr/share/fonts/liberation/LiberationSans-Bold.ttf',
    'C:/Windows/Fonts/arialbd.ttf',
    'C:/Windows/Fonts/calibrib.ttf',
    '/Library/Fonts/Arial Bold.ttf',
]


def _register_pdf_fonts():
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont

    reg_name, bold_name = 'Helvetica', 'Helvetica-Bold'
    for path in _FONT_REGULAR_PATHS:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont('PPE-Regular', path))
                reg_name = 'PPE-Regular'
                break
            except Exception:
                continue
    for path in _FONT_BOLD_PATHS:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont('PPE-Bold', path))
                bold_name = 'PPE-Bold'
                break
            except Exception:
                continue
    return reg_name, bold_name


# ── Data helpers ──────────────────────────────────────────────────────────────

def _fetch_incidents(date_from=None, date_to=None):
    if not os.path.exists(INC_DB):
        return []
    conn = sqlite3.connect(INC_DB)
    conn.row_factory = sqlite3.Row
    q = "SELECT * FROM incidents WHERE 1=1"
    params = []
    if date_from:
        q += " AND created_at >= ?"; params.append(date_from)
    if date_to:
        q += " AND created_at <= ?"; params.append(date_to + 'T23:59:59')
    q += " ORDER BY created_at DESC"
    rows = conn.execute(q, params).fetchall()
    conn.close()
    return [dict(r) for r in rows]


def _fmt_violations(inc, t: dict) -> str:
    raw = inc.get('violation_types', '')
    labels = [t['violations'].get(v.strip(), v.strip()) for v in raw.split(',') if v.strip()]
    return ', '.join(labels) or raw


def _fmt_status(status: str, t: dict) -> str:
    return {
        'new':       t['status_new'],
        'reviewing': t['status_review'],
        'closed':    t['status_closed'],
    }.get(status, status)


# ── PDF ───────────────────────────────────────────────────────────────────────

@router.get("/pdf", dependencies=[require_role("supervisor")])
def generate_pdf(
    date_from: str = Query(None),
    date_to:   str = Query(None),
    title:     str = Query(None),
    lang:      str = Query('pl'),
):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    t = _tr(lang)
    font_reg, font_bold = _register_pdf_fonts()
    report_title = title or t['title_default']

    incidents = _fetch_incidents(date_from, date_to)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=2*cm, rightMargin=2*cm,
                            topMargin=2*cm, bottomMargin=2*cm)

    style_title  = ParagraphStyle('T', fontName=font_bold, fontSize=16, spaceAfter=6)
    style_normal = ParagraphStyle('N', fontName=font_reg,  fontSize=9,  spaceAfter=4)

    story = []
    story.append(Paragraph(report_title, style_title))

    now = datetime.now().strftime('%d.%m.%Y %H:%M')
    pf  = date_from or t['all']
    pt  = date_to   or t['all']
    story.append(Paragraph(f"{t['generated']}: {now}   |   {t['period']}: {pf} – {pt}", style_normal))
    story.append(Spacer(1, 0.4*cm))

    total    = len(incidents)
    new_c    = sum(1 for i in incidents if i.get('status') == 'new')
    closed_c = sum(1 for i in incidents if i.get('status') == 'closed')
    story.append(Paragraph(
        f"{t['summary_total']}: <b>{total}</b>  |  "
        f"{t['summary_new']}: <b>{new_c}</b>  |  "
        f"{t['summary_closed']}: <b>{closed_c}</b>",
        style_normal))
    story.append(Spacer(1, 0.4*cm))

    if incidents:
        headers = [t['col_id'], t['col_date'], t['col_violation'],
                   t['col_zone'], t['col_status'], t['col_track']]
        data = [headers]
        for inc in incidents:
            data.append([
                str(inc['id']),
                inc.get('created_at', '')[:16].replace('T', ' '),
                _fmt_violations(inc, t),
                inc.get('zone_name') or '—',
                _fmt_status(inc.get('status', ''), t),
                str(inc.get('track_id', '')),
            ])

        col_widths = [1.2*cm, 3.5*cm, 5.5*cm, 3.5*cm, 2.5*cm, 2*cm]
        tbl = Table(data, colWidths=col_widths, repeatRows=1)
        tbl.setStyle(TableStyle([
            ('BACKGROUND',    (0, 0), (-1, 0),  colors.HexColor('#1e3a5f')),
            ('TEXTCOLOR',     (0, 0), (-1, 0),  colors.white),
            ('FONTNAME',      (0, 0), (-1, 0),  font_bold),
            ('FONTNAME',      (0, 1), (-1, -1), font_reg),
            ('FONTSIZE',      (0, 0), (-1, -1), 8),
            ('ROWBACKGROUNDS',(0, 1), (-1, -1), [colors.white, colors.HexColor('#f3f4f6')]),
            ('GRID',          (0, 0), (-1, -1), 0.5, colors.HexColor('#d1d5db')),
            ('VALIGN',        (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING',    (0, 0), (-1, -1), 3),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ]))
        story.append(tbl)
    else:
        story.append(Paragraph(t['no_incidents'], style_normal))

    doc.build(story)
    buf.seek(0)
    filename = f"raport_ppe_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    return StreamingResponse(buf, media_type='application/pdf',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'})


# ── Excel ─────────────────────────────────────────────────────────────────────

@router.get("/excel", dependencies=[require_role("supervisor")])
def generate_excel(
    date_from: str = Query(None),
    date_to:   str = Query(None),
    lang:      str = Query('pl'),
):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    t = _tr(lang)
    incidents = _fetch_incidents(date_from, date_to)
    wb = Workbook()

    # ── Incidents sheet ───────────────────────────────────────────────────
    ws = wb.active
    ws.title = t['sheet_incidents']

    headers = [t['col_id'], t['col_date'], t['col_violation'],
               t['col_zone'], t['col_status'], t['col_track'], t['col_notes']]

    hdr_fill = PatternFill(start_color='1e3a5f', end_color='1e3a5f', fill_type='solid')
    hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=10)
    dat_font = Font(name='Calibri', size=9)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    alt_fill = PatternFill(start_color='f3f4f6', end_color='f3f4f6', fill_type='solid')
    for row_idx, inc in enumerate(incidents, 2):
        row_data = [
            inc['id'],
            inc.get('created_at', '')[:16].replace('T', ' '),
            _fmt_violations(inc, t),
            inc.get('zone_name') or '',
            _fmt_status(inc.get('status', ''), t),
            inc.get('track_id', ''),
            inc.get('notes', '') or '',
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = dat_font
            if row_idx % 2 == 0:
                cell.fill = alt_fill

    for col, w in enumerate([8, 18, 35, 20, 16, 10, 30], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20

    # ── Summary sheet ─────────────────────────────────────────────────────
    ws2 = wb.create_sheet(t['sheet_summary'])
    lbl_font = Font(name='Calibri', bold=True, size=10)
    val_font = Font(name='Calibri', size=10)

    now_str = datetime.now().strftime('%d.%m.%Y %H:%M')
    summary = [
        (t['generated'] + ':',      now_str),
        (t['summary_total'] + ':',  len(incidents)),
        (t['summary_new'] + ':',    sum(1 for i in incidents if i.get('status') == 'new')),
        (t['summary_review'] + ':', sum(1 for i in incidents if i.get('status') == 'reviewing')),
        (t['summary_closed'] + ':', sum(1 for i in incidents if i.get('status') == 'closed')),
    ]
    for r, (label, value) in enumerate(summary, 1):
        ws2.cell(row=r, column=1, value=label).font = lbl_font
        ws2.cell(row=r, column=2, value=value).font = val_font
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"raport_ppe_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return StreamingResponse(buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment; filename="{filename}"'})
